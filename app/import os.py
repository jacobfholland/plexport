import os
import csv
import logging
from app.agent import generate_link
from openpyxl import Workbook


class Export:
    def __init__(self, plex, export_dir) -> None:
        self.plex = plex
        self.export_dir = export_dir

    def export(self, format="csv"):
        if format not in ["csv", "xlsx"]:
            raise ValueError("Unsupported format. Choose 'csv' or 'xlsx'.")

        SECTIONS = os.environ.get("SECTIONS", "all_sections").split(",")

        def is_section_valid(section_title):
            return "all_sections" in SECTIONS or section_title in SECTIONS

        def get_filename(section_title):
            return os.path.join(self.export_dir, f"{section_title}.{format}")

        handlers = {
            'movie': (getattr(self, f'write_movies_to_{format}'), f"Processing library section: {{section.title}}"),
            'show': (getattr(self, f'write_tvshows_to_{format}'), f"Processing library section: {{section.title}}"),
            'artist': (getattr(self, f'write_music_to_{format}'), f"Processing library section: {{section.title}}")
        }

        try:
            for section in self.plex.library.sections():
                if is_section_valid(section.title) and section.type in handlers:
                    logging.info(handlers[section.type]
                                 [1].format(section=section))
                    items = section.all()
                    filename = get_filename(section.title)
                    handlers[section.type][0](items, filename)
        except KeyboardInterrupt:
            logging.warning(
                f"Ctrl + C User interrupt. Shutting down gracefully...")
        except Exception as e:
            logging.warning(f"Failed to export to {format.upper()}: {e}")

    def write_movies_to_csv(self, movies, filename):
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Duration (mins)', 'Genres', 'Directors', 'Writers',
                       'Actors', 'Resolution', 'Codec', 'Container', 'Bitrate', 'Size (MB)', 'Audio Channels', 'Added At', 'Updated At']
            writer.writerow(headers)
            for movie in movies:
                movie_data = self.write_movie(movie)
                writer.writerow(self.values_in_order(movie_data))
                logging.info(
                    f"Movie Exported: {movie_data.get('title')} ({movie_data.get('year')})")

    def write_tvshows_to_csv(self, shows, filename):
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Genres',
                       'Total Seasons', 'Total Episodes', 'Added At', 'Updated At']
            writer.writerow(headers)
            for show in shows:
                link = generate_link(show)
                genres = ', '.join(
                    [genre.tag for genre in show.genres]) if show.genres else None
                added_at = show.addedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if show.addedAt else None
                updated_at = show.updatedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if show.updatedAt else None
                writer.writerow([show.title, show.year, link, show.rating, show.summary, genres, len(
                    show.seasons()), len(show.episodes()), added_at, updated_at])
                logging.info(f"TV Show Exported: {show.title} ({show.year})")

    def write_music_to_csv(self, artists, filename):
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            headers = ['Artist Name', 'Genres', 'Albums',
                       'Tracks', 'Added At', 'Updated At']
            writer.writerow(headers)
            for artist in artists:
                genres = ', '.join(
                    [genre.tag for genre in artist.genres]) if artist.genres else None
                added_at = artist.addedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if artist.addedAt else None
                updated_at = artist.updatedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if artist.updatedAt else None
                albums = len(artist.albums())
                tracks = sum(len(album.tracks()) for album in artist.albums())
                writer.writerow([artist.title, genres, albums,
                                 tracks, added_at, updated_at])
                logging.info(f"Music Artist Exported: {artist.title}")

    def write_movies_to_xlsx(self, movies, filename):
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Duration (mins)', 'Genres', 'Directors', 'Writers',
                   'Actors', 'Resolution', 'Codec', 'Container', 'Bitrate', 'Size (MB)', 'Audio Channels', 'Added At', 'Updated At']
        ws.append(headers)
        for movie in movies:
            movie_data = self.write_movie(movie)
            ws.append(self.values_in_order(movie_data))
            logging.info(
                f"Movie Exported: {movie_data.get('title')} ({movie_data.get('year')})")
        wb.save(filename)

    def write_tvshows_to_xlsx(self, shows, filename):
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Genres',
                   'Total Seasons', 'Total Episodes', 'Added At', 'Updated At']
        ws.append(headers)
        for show in shows:
            link = generate_link(show)
            genres = ', '.join([genre.tag for genre in show.genres])
            added_at = show.addedAt.strftime("%Y-%m-%d %H:%M:%S")
            updated_at = show.updatedAt.strftime("%Y-%m-%d %H:%M:%S")
            ws.append([show.title, show.year, link, show.rating, show.summary, genres, len(
                show.seasons()), len(show.episodes()), added_at, updated_at])
            logging.info(f"TV Show Exported: {show.title} ({show.year})")

    def write_music_to_xlsx(self, artists, filename):
        wb = Workbook()
        ws = wb.active
        headers = ['Artist Name', 'Genres', 'Albums',
                   'Tracks', 'Added At', 'Updated At']
        ws.append(headers)
        for artist in artists:
            genres = ', '.join([genre.tag for genre in artist.genres])
            albums = len(artist.albums())
            tracks = sum([len(album.tracks()) for album in artist.albums()])
            added_at = artist.addedAt.strftime("%Y-%m-%d %H:%M:%S")
            updated_at = artist.updatedAt.strftime("%Y-%m-%d %H:%M:%S")
            ws.append([artist.title, genres, albums,
                      tracks, added_at, updated_at])
            logging.info(f"Music Artist Exported: {artist.title}")

    def values_in_order(self, data):
        # This function assumes data is a dict containing movie details.
        # We can use it to order our data correctly for CSV and Excel rows.
        return [data.get(header) for header in ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Duration (mins)', 'Genres', 'Directors', 'Writers',
                                                'Actors', 'Resolution', 'Codec', 'Container', 'Bitrate', 'Size (MB)', 'Audio Channels', 'Added At', 'Updated At']]

    def write_movie(self, movie):
        # This function returns a dictionary of movie details.
        link = generate_link(movie)
        genres = ', '.join([genre.tag for genre in movie.genres])
        directors = ', '.join([director.tag for director in movie.directors])
        writers = ', '.join([writer.tag for writer in movie.writers])
        actors = ', '.join([actor.tag for actor in movie.actors])
        resolution = movie.media[0].videoResolution
        codec = movie.media[0].videoCodec
        container = movie.media[0].container
        bitrate = movie.media[0].bitrate
        size_mb = movie.media[0].size / (1024 * 1024)
        audio_channels = movie.media[0].audioChannels
        added_at = movie.addedAt.strftime("%Y-%m-%d %H:%M:%S")
        updated_at = movie.updatedAt.strftime("%Y-%m-%d %H:%M:%S")
        duration_mins = movie.duration // 60000  # Convert from ms to mins

        movie_data = {
            'Name': movie.title,
            'Year': movie.year,
            'Link': link,
            'Rating': movie.rating,
            'Summary': movie.summary,
            'Duration (mins)': duration_mins,
            'Genres': genres,
            'Directors': directors,
            'Writers': writers,
            'Actors': actors,
            'Resolution': resolution,
            'Codec': codec,
            'Container': container,
            'Bitrate': bitrate,
            'Size (MB)': round(size_mb, 2),
            'Audio Channels': audio_channels,
            'Added At': added_at,
            'Updated At': updated_at
        }
        return movie_data
