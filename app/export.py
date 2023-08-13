import os
import csv
import logging
from app.agent import generate_link
from openpyxl import Workbook


class Export:
    def __init__(self, plex, export_dir) -> None:
        self.plex = plex
        self.export_dir = export_dir

    def csv(self):
        try:
            for section in self.plex.library.sections():
                if section.type == 'movie':
                    logging.info(
                        f"Processing library section: {section.title}")
                    movies = section.all()
                    filename = os.path.join(
                        self.export_dir, f"{section.title}.csv")
                    self.write_movies_to_csv(movies, filename)
                elif section.type == 'show':
                    logging.info(
                        f"Processing library section: {section.title}")
                    shows = section.all()
                    filename = os.path.join(
                        self.export_dir, f"{section.title}.csv")
                    self.write_tvshows_to_csv(shows, filename)
                elif section.type == 'artist':
                    logging.info(
                        f"Processing library section: {section.title}")
                    artists = section.all()
                    filename = os.path.join(
                        self.export_dir, f"{section.title}.csv")
                    self.write_music_to_csv(artists, filename)
        except KeyboardInterrupt:
            logging.warning(
                "Ctrl + C User interrupt. Shutting down gracefully...")
        # except Exception as e:
        #     logging.warning(f"Failed to export to CSV: {e}")

    def write_movie(self, movie):
        if movie.guid is None or movie.guid.startswith("local://"):
            logging.warning(f"Unmatched movie detected: {movie.title}")
        link = generate_link(movie)
        duration_mins = round(
            movie.duration / 60000) if movie.duration else None
        genres = ', '.join(
            [genre.tag for genre in movie.genres]) if movie.genres else None
        directors = ', '.join(
            [director.tag for director in movie.directors]) if movie.directors else None
        writers = ', '.join(
            [writer.tag for writer in movie.writers]) if movie.writers else None
        actors = ', '.join(
            [actor.tag for actor in movie.actors]) if movie.actors else None

        if movie.media:
            media = movie.media[0]
            resolution = media.videoResolution
            codec = media.videoCodec
            container = media.container
            bitrate = media.bitrate
            size_MB = media.size / \
                (1024 * 1024) if hasattr(media, 'size') else None
            audio_channels = media.audioChannels if media.audioChannels else None
        else:
            resolution, codec, container, bitrate, size_MB, audio_channels = [
                None] * 6
            logging.warning(f"Movie {movie.title} has no size")

        added_at = movie.addedAt.strftime(
            '%Y-%m-%d %H:%M:%S') if movie.addedAt else None
        updated_at = movie.updatedAt.strftime(
            '%Y-%m-%d %H:%M:%S') if movie.updatedAt else None
        return {
            "title": movie.title,
            "year": movie.year,
            "link": link,
            "rating": movie.rating,
            "summary": movie.summary,
            "duration_mins": duration_mins,
            "genres": genres,
            "directors": directors,
            "writers": writers,
            "actors": actors,
            "resolution": resolution,
            "codec": codec,
            "container": container,
            "bitrate": bitrate,
            "size_MB": size_MB,
            "audio_channels": audio_channels,
            "added_at": added_at,
            "updated_at": updated_at
        }

    def values_in_order(self, dictionary):
        return list(dictionary.values())

    def write_movies_to_csv(self, movies, filename):
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Duration (mins)', 'Genres', 'Directors', 'Writers',
                       'Actors', 'Resolution', 'Codec', 'Container', 'Bitrate', 'Size (MB)', 'Audio Channels', 'Added At', 'Updated At']
            writer.writerow(headers)
            for movie in movies:
                if not movie.title:
                    print(movie)
                movie = self.write_movie(movie)
                # try:
                #     logging.info(
                #         f"Movie Exported: {movie.get('title')} ({movie.get('year')})")
                # except Exception:
                #     print(type(movie), vars(movie))

                writer.writerow(self.values_in_order(movie))
                logging.info(
                    f"Movie Exported: {movie.get('title')} ({movie.get('year')})")

    def write_tvshows_to_csv(self, shows, filename):
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Genres',
                       'Total Seasons', 'Total Episodes', 'Added At', 'Updated At']
            writer.writerow(headers)
            for show in shows:
                link = generate_link(show)
                genres = ', '.join(
                    [genre.tag for genre in show.genres]) if show.genres else None
                added_at = show.addedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if show.addedAt else None
                updated_at = show.updatedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if show.updatedAt else None
                writer.writerow([show.title, show.year, link, show.rating, show.summary, genres, len(
                    show.seasons()), len(show.episodes()), added_at, updated_at])
                logging.info(f"TV Show Exported: {show.title} ({show.year})")

    def write_music_to_csv(self, artists, filename):
        with open(filename, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            headers = ['Artist Name', 'Genres', 'Albums',
                       'Tracks', 'Added At', 'Updated At']
            writer.writerow(headers)
            for artist in artists:
                genres = ', '.join(
                    [genre.tag for genre in artist.genres]) if artist.genres else None
                added_at = artist.addedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if artist.addedAt else None
                updated_at = artist.updatedAt.strftime(
                    '%Y-%m-%d %H:%M:%S') if artist.updatedAt else None
                albums = len(artist.albums())
                tracks = sum(len(album.tracks()) for album in artist.albums())
                writer.writerow([artist.title, genres, albums,
                                tracks, added_at, updated_at])
                logging.info(f"Music Artist Exported: {artist.title}")

    def xlsx(self):
        try:
            for section in self.plex.library.sections():
                if section.type == 'movie':
                    logging.info(
                        f"Processing library section: {section.title}")
                    movies = section.all()
                    filename = os.path.join(
                        self.export_dir, f"{section.title}.xlsx")
                    self.write_movies_to_xlsx(movies, filename)
                elif section.type == 'show':
                    logging.info(
                        f"Processing library section: {section.title}")
                    shows = section.all()
                    filename = os.path.join(
                        self.export_dir, f"{section.title}.xlsx")
                    self.write_tvshows_to_xlsx(shows, filename)
                elif section.type == 'artist':
                    logging.info(
                        f"Processing library section: {section.title}")
                    artists = section.all()
                    filename = os.path.join(
                        self.export_dir, f"{section.title}.xlsx")
                    self.write_music_to_xlsx(artists, filename)
        except KeyboardInterrupt:
            logging.warning(
                "Ctrl + C User interrupt. Shutting down gracefully...")
        except Exception as e:
            logging.warning(f"Failed to export to XLSX: {e}")

    def write_movies_to_xlsx(self, movies, filename):
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Duration (mins)', 'Genres', 'Directors', 'Writers',
                   'Actors', 'Resolution', 'Codec', 'Container', 'Bitrate', 'Size (MB)', 'Audio Channels', 'Added At', 'Updated At']
        ws.append(headers)
        for movie in movies:
            # ... [Existing logic to extract movie attributes]

            ws.append([movie.title, movie.year, link, movie.rating, movie.summary, duration_mins, genres, directors,
                       writers, actors, resolution, codec, container, bitrate, size_MB, audio_channels, added_at, updated_at])
            logging.info(f"Movie Exported: {movie.title} ({movie.year})")
        wb.save(filename)

    def write_tvshows_to_xlsx(self, shows, filename):
        wb = Workbook()
        ws = wb.active
        headers = ['Name', 'Year', 'Link', 'Rating', 'Summary', 'Genres',
                   'Total Seasons', 'Total Episodes', 'Added At', 'Updated At']
        ws.append(headers)
        for show in shows:
            # ... [Existing logic to extract show attributes]

            ws.append([show.title, show.year, link, show.rating, show.summary, genres, len(
                show.seasons()), len(show.episodes()), added_at, updated_at])
            logging.info(f"TV Show Exported: {show.title} ({show.year})")
        wb.save(filename)

    def write_music_to_xlsx(self, artists, filename):
        wb = Workbook()
        ws = wb.active
        headers = ['Artist Name', 'Genres', 'Albums',
                   'Tracks', 'Added At', 'Updated At']
        ws.append(headers)
        for artist in artists:
            # ... [Existing logic to extract artist attributes]

            ws.append([artist.title, genres, albums,
                      tracks, added_at, updated_at])
            logging.info(f"Music Artist Exported: {artist.title}")
        wb.save(filename)
